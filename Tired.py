import tkinter as tk
import tkinter.ttk as ttk
import openpyxl
from tkinter import ttk
from openpyxl.utils import get_column_letter
from tkinter import simpledialog

class ExcelViewer(tk.Frame):
    def __init__(self, parent, path):
        super().__init__(parent)
        self.path = path
        self.workbook = openpyxl.load_workbook(self.path)
        self.sheet = self.workbook.active
        self.row_count = self.sheet.max_row
        self.column_count = self.sheet.max_column
        self.selected_rows = []

        # Create treeview
        self.tree = ttk.Treeview(self, columns=list(range(1, self.column_count+1)), show='headings')
        for i in range(1, self.column_count+1):
            column_name = self.sheet.cell(row=1, column=i).value
            self.tree.heading(i, text=column_name)
        for i in range(2, self.row_count+1):
            row_data = [self.sheet.cell(row=i, column=j).value for j in range(1, self.column_count+1)]
            self.tree.insert('', 'end', values=row_data)
        self.tree.pack(side='left', fill='both', expand=True)

        
       

        # Create button to add new record
        button = ttk.Button(self, text='Add Record', command=self.add_record)
        button.pack(side='top', pady=5)
        button = ttk.Button(parent, text= "Add Record", command=self.add_record)
        button.pack(padx=5, pady=5, side=tk.TOP)
        button = ttk.Button(parent, text= "Delete Record", command=self.select_row)
        button.pack(padx=5, pady=5, side=tk.RIGHT)

        # Add scrollbar
        yscrollbar = ttk.Scrollbar(self, orient='vertical', command=self.tree.yview)
        yscrollbar.pack(side='right', fill='y')
        self.tree.configure(yscrollcommand=yscrollbar.set)

        self.tree.bind('<Button-1>', self.select_row) # moved inside the __init__ method
 
    def select_row(self, event):
        # Clear previously selected rows
        for row_id in self.selected_rows:
            self.tree.item(row_id, tags=())
        self.selected_rows.clear()

        # Select the clicked row
        row_id = self.tree.identify_row(event.y)
        self.tree.item(row_id, tags=('selected',))
        self.selected_rows.append(row_id)


    def add_record(self):
        # Create a new window for adding a record
        add_record_window = tk.Toplevel(self)
        add_record_window.title("Add Record")
        
        # Create a frame to hold the entry widgets
        entry_frame = tk.Frame(add_record_window)
        entry_frame.pack(side=tk.TOP, padx=5, pady=5)
        
        # Add the entry widgets
        parent_label = tk.Label(entry_frame, text="Parent")
        parent_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        parent_entry = tk.Entry(entry_frame)
        parent_entry.grid(row=0, column=1, padx=5, pady=5)
        
        name_label = tk.Label(entry_frame, text="Name")
        name_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        name_entry = tk.Entry(entry_frame)
        name_entry.grid(row=1, column=1, padx=5, pady=5)
        
        address_label = tk.Label(entry_frame, text="Address")
        address_label.grid(row=2, column=2, padx=5,pady=5, sticky=tk.W)
        address_entry = tk.Entry(entry_frame)
        address_entry.grid(row=1, column=1, padx=5, pady=5)

        parent = simpledialog.askstring("Input", "Enter Parent, or leave blank if none")
        name = simpledialog.askstring("Input", "What is customer Name?:")
        address = simpledialog.askstring("Input", "What is customer Address?: ")
        company = simpledialog.askstring("Input", "What is name of Company?: ")
        dateOfSale = simpledialog.askstring("Input", "What was the Date of Sale?: ")
        cost = simpledialog.askstring("Input", "What is Costs of Services?: ")
        services = simpledialog.askstring("Input", "What is Total Services?: ")
        totalService = simpledialog.askstring("Input", "What does the Customer owe?: ")
        amountOwed = simpledialog.askstring("Input", "Enter Costs")
        self.tree.insert(parent, tk.END, text= name, values = (name,address, company, dateOfSale, cost, services, totalService, amountOwed,))

if __name__ == '__main__':
    path = 'C:\\Users\\roshe\\Hello_World\\Project\\Customer Moats- Inc.xlsx'
    root = tk.Tk()
    root.title('MOATS INC')
    # Create background image and place treeview on it
    bg_image = tk.PhotoImage(file="C:\\Users\\roshe\\Hello_World\\Project\\pltp5ptug02ov9qqsqju82sfro.png")
    bg_label = tk.Label(root, image=bg_image)
    bg_label.place(x=0, y=0, relwidth=1, relheight=1)
    ExcelViewer(root, path).pack(side='top', fill='both', expand=True)
    root.mainloop()